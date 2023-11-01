from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
wb = Workbook()
ws = wb.active
ws.title = "Python"
for i in range(13):
  a = ws.cell(row=i+1, column=1, value=i+3)
  if (i+1)%3 ==0:
    b = ws.cell(row=i+1, column=2, value=(i+3)*2)
  else:
    b = ws.cell(row=i+1, column=2, value=i+3)
values = Reference(worksheet=ws,
                 min_row=1,
                 max_row=13,
                 min_col=2,
                 max_col=2)
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.type = "col"
chart.style = 10
chart.title = "Столбчатая диаграмма"
chart.y_axis.title = 'Размер значений столбца B'
chart.y_axis.delete = False
chart.x_axis.title = 'Значения столбца D'
chart.x_axis.delete = False
ws.add_chart(chart, "D2")
chart.width = 10
chart.height = 6
chart2 = PieChart()
chart2.add_data(values, titles_from_data=False)
chart2.title = "Круговая диаграмма"
chart2.width = 10
chart2.height = 10
ws.add_chart(chart2, "B15")
wb.save('test.py.xlsx')